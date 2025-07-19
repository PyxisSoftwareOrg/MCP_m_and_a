"""
Export and reporting tools for M&A Research Assistant
"""

import csv
import io
import json
import logging
import time
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..services import S3Service

logger = logging.getLogger(__name__)

# Initialize services
s3_service = S3Service()


async def export_report(
    company_names: List[str],
    format: str = "json",
    include_raw_data: bool = False
) -> Dict[str, Any]:
    """Generate formatted reports for companies"""
    try:
        logger.info(f"Exporting report for {len(company_names)} companies in {format} format")
        
        if format not in ["json", "csv", "excel"]:
            return {
                "success": False,
                "error": f"Unsupported format: {format}. Use json, csv, or excel"
            }
        
        # Collect company data
        companies_data = []
        missing_companies = []
        
        for company_name in company_names:
            analysis = await s3_service.get_analysis_result(company_name)
            if analysis:
                company_data = {
                    "company_name": analysis.company_name,
                    "website_url": analysis.website_url,
                    "linkedin_url": analysis.linkedin_url,
                    "overall_score": analysis.overall_score,
                    "automated_tier": analysis.automated_tier,
                    "effective_tier": analysis.effective_tier,
                    "is_qualified": analysis.qualification_result.is_qualified,
                    "qualification_score": analysis.qualification_result.qualification_score,
                    "analysis_timestamp": analysis.analysis_timestamp,
                    "recommendation": analysis.recommendation,
                    "key_strengths": analysis.key_strengths,
                    "key_concerns": analysis.key_concerns
                }
                
                # Add dimension scores
                for dim_id, dim_score in analysis.default_scores.items():
                    company_data[f"score_{dim_id}"] = dim_score.score
                    company_data[f"confidence_{dim_id}"] = dim_score.confidence
                
                # Add filtering results
                if analysis.filtering_result:
                    company_data["geographic_region"] = analysis.filtering_result.geographic_region
                    company_data["business_model_type"] = analysis.filtering_result.business_model_type
                    company_data["estimated_revenue"] = analysis.filtering_result.estimated_revenue
                    company_data["estimated_employees"] = analysis.filtering_result.estimated_employees
                    company_data["company_age_years"] = analysis.filtering_result.company_age_years
                
                # Add investment thesis if available
                if analysis.investment_thesis:
                    company_data["investment_recommendation"] = analysis.investment_thesis.recommendation
                    company_data["vms_alignment_score"] = analysis.investment_thesis.vms_alignment_score
                    company_data["confidence_level"] = analysis.investment_thesis.confidence_level
                
                if include_raw_data:
                    company_data["raw_analysis"] = analysis.dict()
                
                companies_data.append(company_data)
            else:
                missing_companies.append(company_name)
        
        if not companies_data:
            return {
                "success": False,
                "error": "No analysis data found for any of the specified companies"
            }
        
        # Generate report based on format
        if format == "json":
            report_content = {
                "export_metadata": {
                    "export_timestamp": datetime.utcnow().isoformat() + 'Z',
                    "format": "json",
                    "companies_requested": len(company_names),
                    "companies_found": len(companies_data),
                    "missing_companies": missing_companies,
                    "include_raw_data": include_raw_data
                },
                "companies": companies_data
            }
            
            # Save to S3
            export_id = f"export_{int(time.time())}"
            s3_key = f"exports/{datetime.now().strftime('%Y-%m-%d')}/{export_id}.json"
            
            await s3_service._save_json_object(s3_key, report_content)
            
            # Generate presigned URL
            presigned_url = await s3_service.generate_presigned_url(s3_key, expiration=3600)
            
            return {
                "success": True,
                "format": "json",
                "export_id": export_id,
                "s3_path": f"s3://{s3_service.bucket_name}/{s3_key}",
                "presigned_url": presigned_url,
                "companies_exported": len(companies_data),
                "missing_companies": missing_companies
            }
        
        elif format == "csv":
            # Convert to CSV
            if not companies_data:
                return {
                    "success": False,
                    "error": "No data to export"
                }
            
            csv_buffer = io.StringIO()
            
            # Get all field names
            fieldnames = set()
            for company in companies_data:
                fieldnames.update(company.keys())
            fieldnames = sorted(list(fieldnames))
            
            writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
            writer.writeheader()
            
            for company in companies_data:
                # Convert lists to strings for CSV
                row = {}
                for key, value in company.items():
                    if isinstance(value, list):
                        row[key] = "; ".join(str(item) for item in value)
                    else:
                        row[key] = value
                writer.writerow(row)
            
            csv_content = csv_buffer.getvalue()
            
            # Save to S3
            export_id = f"export_{int(time.time())}"
            s3_key = f"exports/{datetime.now().strftime('%Y-%m-%d')}/{export_id}.csv"
            
            await s3_service.s3_client.put_object(
                Bucket=s3_service.bucket_name,
                Key=s3_key,
                Body=csv_content.encode('utf-8'),
                ContentType='text/csv',
                ServerSideEncryption='AES256'
            )
            
            # Generate presigned URL
            presigned_url = await s3_service.generate_presigned_url(s3_key, expiration=3600)
            
            return {
                "success": True,
                "format": "csv",
                "export_id": export_id,
                "s3_path": f"s3://{s3_service.bucket_name}/{s3_key}",
                "presigned_url": presigned_url,
                "companies_exported": len(companies_data),
                "missing_companies": missing_companies
            }
        
        elif format == "excel":
            # Use the XLSX generation function
            return await generate_xlsx_export(
                companies=company_names,
                include_charts=True,
                custom_fields=[]
            )
        
    except Exception as e:
        logger.error(f"Error exporting report: {e}")
        return {
            "success": False,
            "error": str(e)
        }


async def generate_xlsx_export(
    companies: List[str],
    include_charts: bool = True,
    custom_fields: List[str] = None
) -> Dict[str, Any]:
    """Generate downloadable XLSX files with formatting"""
    try:
        logger.info(f"Generating XLSX export for {len(companies)} companies")
        
        # Collect data for all companies
        companies_data = []
        dimension_data = []
        summary_data = []
        
        for company_name in companies:
            analysis = await s3_service.get_analysis_result(company_name)
            if not analysis:
                continue
            
            # Main company data
            company_row = {
                "Company Name": analysis.company_name,
                "Website": analysis.website_url,
                "LinkedIn": analysis.linkedin_url or "",
                "Overall Score": analysis.overall_score,
                "Automated Tier": analysis.automated_tier,
                "Effective Tier": analysis.effective_tier,
                "Qualified": "Yes" if analysis.qualification_result.is_qualified else "No",
                "Qualification Score": analysis.qualification_result.qualification_score,
                "Analysis Date": analysis.analysis_timestamp[:10],  # Date only
                "Recommendation": analysis.recommendation[:100] + "..." if len(analysis.recommendation) > 100 else analysis.recommendation
            }
            
            # Add filtering data
            if analysis.filtering_result:
                company_row["Geographic Region"] = analysis.filtering_result.geographic_region
                company_row["Business Model"] = analysis.filtering_result.business_model_type
                company_row["Est. Revenue"] = analysis.filtering_result.estimated_revenue
                company_row["Est. Employees"] = analysis.filtering_result.estimated_employees
                company_row["Company Age"] = analysis.filtering_result.company_age_years
            
            # Add investment thesis data
            if analysis.investment_thesis:
                company_row["Investment Rec"] = analysis.investment_thesis.recommendation
                company_row["VMS Alignment"] = analysis.investment_thesis.vms_alignment_score
                company_row["Confidence"] = analysis.investment_thesis.confidence_level
            
            companies_data.append(company_row)
            
            # Dimension scores for separate sheet
            for dim_id, dim_score in analysis.default_scores.items():
                dimension_data.append({
                    "Company Name": analysis.company_name,
                    "Dimension": dim_id.replace("_", " ").title(),
                    "Score": dim_score.score,
                    "Confidence": dim_score.confidence,
                    "Evidence Count": len(dim_score.evidence),
                    "Top Evidence": dim_score.evidence[0] if dim_score.evidence else ""
                })
        
        if not companies_data:
            return {
                "success": False,
                "error": "No analysis data found for specified companies"
            }
        
        # Create summary statistics
        scores = [row["Overall Score"] for row in companies_data]
        qualified_count = sum(1 for row in companies_data if row["Qualified"] == "Yes")
        
        summary_data = [
            {"Metric": "Total Companies", "Value": len(companies_data)},
            {"Metric": "Qualified Companies", "Value": qualified_count},
            {"Metric": "Qualification Rate", "Value": f"{qualified_count/len(companies_data)*100:.1f}%"},
            {"Metric": "Average Score", "Value": f"{sum(scores)/len(scores):.2f}"},
            {"Metric": "Highest Score", "Value": f"{max(scores):.2f}"},
            {"Metric": "Lowest Score", "Value": f"{min(scores):.2f}"},
        ]
        
        # Tier distribution
        tier_counts = {}
        for row in companies_data:
            tier = row["Effective Tier"]
            tier_counts[tier] = tier_counts.get(tier, 0) + 1
        
        for tier, count in tier_counts.items():
            summary_data.append({
                "Metric": f"{tier} Tier Count",
                "Value": count
            })
        
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        summary_ws = wb.create_sheet("Summary")
        companies_ws = wb.create_sheet("Companies")
        dimensions_ws = wb.create_sheet("Dimension Scores")
        
        # Summary sheet
        summary_df = pd.DataFrame(summary_data)
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            summary_ws.append(r)
        
        # Format summary sheet
        summary_ws['A1'].font = Font(bold=True)
        summary_ws['B1'].font = Font(bold=True)
        for row in summary_ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        
        # Companies sheet
        companies_df = pd.DataFrame(companies_data)
        for r in dataframe_to_rows(companies_df, index=False, header=True):
            companies_ws.append(r)
        
        # Format companies sheet
        for row in companies_ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        
        # Add conditional formatting for scores
        for row_num in range(2, len(companies_data) + 2):
            score_cell = companies_ws[f'D{row_num}']  # Overall Score column
            score_value = score_cell.value
            
            if score_value >= 8.0:
                score_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif score_value >= 6.0:
                score_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
            elif score_value >= 4.0:
                score_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
            else:
                score_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        
        # Dimensions sheet
        if dimension_data:
            dimensions_df = pd.DataFrame(dimension_data)
            for r in dataframe_to_rows(dimensions_df, index=False, header=True):
                dimensions_ws.append(r)
            
            # Format dimensions sheet
            for row in dimensions_ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for ws in [summary_ws, companies_ws, dimensions_ws]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        
        # Save to S3
        export_id = f"xlsx_export_{int(time.time())}"
        s3_key = f"exports/{datetime.now().strftime('%Y-%m-%d')}/xlsx_exports/{export_id}.xlsx"
        
        await s3_service.s3_client.put_object(
            Bucket=s3_service.bucket_name,
            Key=s3_key,
            Body=excel_content,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ServerSideEncryption='AES256'
        )
        
        # Generate presigned URL (24 hour expiration)
        presigned_url = await s3_service.generate_presigned_url(s3_key, expiration=86400)
        
        return {
            "success": True,
            "format": "xlsx",
            "export_id": export_id,
            "s3_path": f"s3://{s3_service.bucket_name}/{s3_key}",
            "presigned_url": presigned_url,
            "companies_exported": len(companies_data),
            "file_size_bytes": len(excel_content),
            "sheets": ["Summary", "Companies", "Dimension Scores"],
            "expiration": (datetime.now() + timedelta(hours=24)).isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error generating XLSX export: {e}")
        return {
            "success": False,
            "error": str(e)
        }